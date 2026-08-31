import json
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

SCRIPT = Path(__file__).parent.parent / "skills" / "excel" / "scripts" / "dough_excel.py"


def run(*argv):
    return subprocess.run([sys.executable, str(SCRIPT), *argv], capture_output=True, text=True)


def csv_path(tmp_path, name, text):
    p = tmp_path / name
    p.write_text(text, encoding="utf-8")
    return str(p)


def payload(tmp_path, entries):
    p = tmp_path / "payload.json"
    p.write_text(json.dumps({"entries": entries}))
    return str(p)


def entry(tmp_path, sheet="Revenue Detail", rows="month,revenue\n2026-01,100\n2026-02,200\n"):
    return {
        "sheet": sheet,
        "queryId": "q-1",
        "queryName": "Revenue by month",
        "sqlSnapshot": "SELECT month, revenue FROM finance.revenue",
        "refreshedAt": "2026-07-25T14:03:00Z",
        "rowCount": 2,
        "csvPath": csv_path(tmp_path, f"{sheet}.csv", rows),
        "refreshNotes": "USD, sorted by month",
    }


def test_create_into_empty(tmp_path):
    book = tmp_path / "book.xlsx"
    result = run("create", str(book), "--payload", payload(tmp_path, [entry(tmp_path)]))
    assert result.returncode == 0, result.stderr
    wb = load_workbook(book)
    assert wb.sheetnames[0] == "Dough"
    assert wb["Dough"]["A2"].value == "dough-manifest v1"
    assert wb["Dough"]["A4"].value == "Revenue Detail"
    assert wb["Dough"]["B4"].value == "q-1"
    assert wb["Dough"]["E4"].value == "2026-07-25T14:03:00Z"
    assert wb["Dough"]["F4"].value == 2  # row_count is its own numeric column
    ws = wb["Revenue Detail"]
    assert "Managed by Dough" in ws["A1"].value
    assert ws["A2"].value == "month"
    assert ws["B3"].value == 100  # numbers coerced from CSV strings


def test_create_into_existing_preserves_other_sheets(tmp_path):
    book = tmp_path / "book.xlsx"
    wb = Workbook()
    wb.active.title = "Model"
    wb.active["A1"] = "=SUM('Revenue Detail'!B:B)"
    wb.save(book)
    result = run("create", str(book), "--payload", payload(tmp_path, [entry(tmp_path)]))
    assert result.returncode == 0, result.stderr
    wb = load_workbook(book)
    assert "Model" in wb.sheetnames
    assert wb["Model"]["A1"].value.startswith("=SUM")
    assert wb.sheetnames[0] == "Dough"


def test_refresh_shrunk_rows_and_mangled_banner_restored(tmp_path):
    book = tmp_path / "book.xlsx"
    run("create", str(book), "--payload", payload(tmp_path, [entry(tmp_path)]))
    wb = load_workbook(book)
    wb["Revenue Detail"]["A1"] = "user typed over the banner"
    wb.save(book)
    shorter = entry(tmp_path, rows="month,revenue\n2026-03,300\n")
    shorter["rowCount"] = 1
    shorter["refreshedAt"] = "2026-08-01T09:00:00Z"
    result = run("refresh", str(book), "--all", "--payload", payload(tmp_path, [shorter]))
    assert result.returncode == 0, result.stderr
    wb = load_workbook(book)
    ws = wb["Revenue Detail"]
    assert "Managed by Dough" in ws["A1"].value  # banner self-healed
    assert "2026-08-01" in ws["A1"].value        # with the new date
    assert ws["A3"].value == "2026-03"
    assert ws["A4"].value is None                # old rows gone
    assert wb["Dough"]["E4"].value == "2026-08-01T09:00:00Z"  # timestamp updated in place
    assert wb["Dough"]["F4"].value == 1                       # row_count updated in place


def test_refresh_unknown_sheet_is_drift(tmp_path):
    book = tmp_path / "book.xlsx"
    run("create", str(book), "--payload", payload(tmp_path, [entry(tmp_path)]))
    rogue = entry(tmp_path, sheet="Not In Manifest")
    result = run("refresh", str(book), "--all", "--payload", payload(tmp_path, [rogue]))
    assert result.returncode == 3
    assert "reconcile" in result.stderr


def test_newer_manifest_version_stops(tmp_path):
    book = tmp_path / "book.xlsx"
    run("create", str(book), "--payload", payload(tmp_path, [entry(tmp_path)]))
    wb = load_workbook(book)
    wb["Dough"]["A2"] = "dough-manifest v2"
    wb.save(book)
    result = run("list", str(book))
    assert result.returncode == 3
    assert "update the Dough plugin" in result.stderr


def test_older_manifest_version_says_migrate_not_update(tmp_path):
    book = tmp_path / "book.xlsx"
    run("create", str(book), "--payload", payload(tmp_path, [entry(tmp_path)]))
    wb = load_workbook(book)
    wb["Dough"]["A2"] = "dough-manifest v0"
    wb.save(book)
    result = run("list", str(book))
    assert result.returncode == 3
    assert "predates" in result.stderr
    assert "update the Dough plugin" not in result.stderr


def test_refresh_preserves_tab_order(tmp_path):
    book = tmp_path / "book.xlsx"
    run("create", str(book), "--payload", payload(tmp_path, [entry(tmp_path)]))
    wb = load_workbook(book)
    wb.create_sheet("ZZ_UserModel")
    wb.save(book)
    before = load_workbook(book).sheetnames
    result = run("refresh", str(book), "--all", "--payload", payload(tmp_path, [entry(tmp_path)]))
    assert result.returncode == 0, result.stderr
    assert load_workbook(book).sheetnames == before


def test_list_outputs_manifest_json(tmp_path):
    book = tmp_path / "book.xlsx"
    run("create", str(book), "--payload", payload(tmp_path, [entry(tmp_path)]))
    result = run("list", str(book))
    assert result.returncode == 0, result.stderr
    parsed = json.loads(result.stdout)
    assert parsed["version"] == "dough-manifest v1"
    assert parsed["entries"][0]["query_id"] == "q-1"
    assert parsed["entries"][0]["sheet"] == "Revenue Detail"
    assert parsed["entries"][0]["row_count"] == "2"


def test_non_managed_workbook_fails_cleanly(tmp_path):
    book = tmp_path / "plain.xlsx"
    Workbook().save(book)
    result = run("list", str(book))
    assert result.returncode == 3
    assert "not a Dough-managed workbook" in result.stderr


def test_row_cap_enforced(tmp_path):
    book = tmp_path / "book.xlsx"
    big = "col\n" + "\n".join(str(i) for i in range(20001)) + "\n"
    capped = entry(tmp_path, rows=big)
    result = run("create", str(book), "--payload", payload(tmp_path, [capped]))
    assert result.returncode == 2
    assert "20000 cap" in result.stderr


def test_rows_just_under_cap_are_written(tmp_path):
    # The cap is a ceiling, not a target: a sheet an order of magnitude past the
    # old 5_000 limit must write cleanly. Guards against the cap silently
    # regressing to the API's page size again.
    book = tmp_path / "book.xlsx"
    rows = "col\n" + "\n".join(str(i) for i in range(19_999)) + "\n"
    result = run("create", str(book), "--payload",
                 payload(tmp_path, [entry(tmp_path, rows=rows)]))
    assert result.returncode == 0, result.stderr
    ws = load_workbook(book)["Revenue Detail"]
    assert ws.max_row == 20_001          # banner + header + 19,999 data rows


def test_identifiers_not_corrupted_by_coercion(tmp_path):
    # Zero-padded GL codes, phone-like strings, exponents, underscores, and
    # >15-digit ids must survive as strings; real decimals become numbers.
    rows = (
        "code,amount\n"
        "0100,15000\n"
        "02134,-42.5\n"
        "+16175551234,0\n"
        "1e5,0.5\n"
        "1_0,999999999999999\n"
        "00012345678901234567890,7\n"
    )
    e = entry(tmp_path, rows=rows)
    e["rowCount"] = 6
    book = tmp_path / "book.xlsx"
    result = run("create", str(book), "--payload", payload(tmp_path, [e]))
    assert result.returncode == 0, result.stderr
    ws = load_workbook(book)["Revenue Detail"]
    codes = [ws.cell(row=r, column=1).value for r in range(3, 9)]
    assert codes == ["0100", "02134", "+16175551234", "1e5", "1_0", "00012345678901234567890"]
    assert all(isinstance(c, str) for c in codes)
    amounts = [ws.cell(row=r, column=2).value for r in range(3, 9)]
    assert amounts == [15000, -42.5, 0, 0.5, 999999999999999, 7]
    assert isinstance(amounts[0], int) and isinstance(amounts[1], float)


def test_missing_payload_fields_rejected(tmp_path):
    book = tmp_path / "book.xlsx"
    bad = entry(tmp_path)
    del bad["queryId"]
    result = run("create", str(book), "--payload", payload(tmp_path, [bad]))
    assert result.returncode == 2
    assert "queryId" in result.stderr
