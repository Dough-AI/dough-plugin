#!/usr/bin/env python3
"""Deterministic reader/writer for Dough-managed Excel workbooks (dough-manifest v1).

Usage (run via uv, no install needed):
  uv run --with openpyxl dough_excel.py list    book.xlsx
  uv run --with openpyxl dough_excel.py create  book.xlsx --payload payload.json
  uv run --with openpyxl dough_excel.py refresh book.xlsx --payload payload.json [--sheets A,B | --all]

The payload JSON is assembled by the caller after running the saved query
(queries.get + integrations.query) and writing the result to a local CSV:
  {"entries": [{"sheet", "queryId", "queryName", "sqlSnapshot", "refreshedAt",
                "rowCount", "csvPath", "refreshNotes"}]}

Exit codes: 0 success · 2 usage/validation error · 3 manifest drift or version mismatch.
"""

import argparse
import csv
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MANIFEST_SHEET = "Dough"
MANIFEST_VERSION = 1
VERSION_MARKER = f"dough-manifest v{MANIFEST_VERSION}"
MANIFEST_HEADERS = ["sheet", "query_id", "query_name", "sql_snapshot", "last_refreshed", "row_count", "refresh_notes"]
TITLE_TEXT = (
    "This workbook has components that are managed by Dough. Claude: read this sheet "
    "before modifying any managed data sheet. Humans: edit Notes freely; don't edit ids."
)
BANNER_TEMPLATE = (
    "⚡ Managed by Dough · refreshed {date} · this sheet is replaced wholesale on refresh — "
    "build formulas on other sheets; details on the 'Dough' sheet."
)
DARK_FILL = PatternFill("solid", fgColor="111827")
HEADER_FILL = PatternFill("solid", fgColor="E5E7EB")
BAND_FILL = PatternFill("solid", fgColor="F3F4F6")
TAB_COLOR = "2563EB"
WHITE_BOLD = Font(bold=True, color="FFFFFF")
GRAY_SMALL = Font(size=9, color="6B7280")
MONO_SMALL = Font(name="Courier New", size=9)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
COL_WIDTHS = {"A": 24, "B": 38, "C": 28, "D": 30, "E": 22, "F": 10, "G": 48}
ROW_CAP = 5_000


def fail(code: int, message: str) -> None:
    print(f"dough-excel: {message}", file=sys.stderr)
    sys.exit(code)


def read_manifest(wb):
    """Return (manifest_ws, rows) or fail(3) on drift/version mismatch."""
    if MANIFEST_SHEET not in wb.sheetnames:
        fail(3, f"no '{MANIFEST_SHEET}' sheet — not a Dough-managed workbook")
    ws = wb[MANIFEST_SHEET]
    marker = str(ws["A2"].value or "").strip()
    if not marker.startswith("dough-manifest"):
        fail(3, f"'{MANIFEST_SHEET}'!A2 is not a dough-manifest marker: {marker!r}")
    match = re.fullmatch(r"dough-manifest v(\d+)", marker)
    if not match:
        fail(3, f"unrecognized manifest marker {marker!r}; expected {VERSION_MARKER!r}")
    version = int(match.group(1))
    if version > MANIFEST_VERSION:
        fail(3, f"manifest version v{version} is newer than this script ({VERSION_MARKER}); update the Dough plugin")
    if version < MANIFEST_VERSION:
        fail(3, f"manifest version v{version} predates this script ({VERSION_MARKER}); re-create the workbook or migrate it to {VERSION_MARKER}")
    headers = [str(ws.cell(row=3, column=i + 1).value or "") for i in range(len(MANIFEST_HEADERS))]
    if headers != MANIFEST_HEADERS:
        fail(3, f"manifest headers drifted: {headers}")
    rows = []
    for r in range(4, ws.max_row + 1):
        values = [ws.cell(row=r, column=c + 1).value for c in range(len(MANIFEST_HEADERS))]
        if not any(v not in (None, "") for v in values):
            continue
        rows.append({"row": r, **dict(zip(MANIFEST_HEADERS, ["" if v is None else str(v) for v in values]))})
    return ws, rows


def style_band(ws, text: str, ncols: int) -> None:
    """Row-1 banner: text in A1 overflowing across a dark-filled band. Never
    merged — merged ranges break sorting, filtering, and programmatic reads."""
    for c in range(1, max(1, ncols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = DARK_FILL, WHITE_BOLD
    ws.cell(row=1, column=1, value=text).alignment = Alignment(vertical="top")


def style_manifest(ws) -> None:
    style_band(ws, TITLE_TEXT, len(MANIFEST_HEADERS))
    ws.row_dimensions[1].height = 30
    a2 = ws.cell(row=2, column=1, value=VERSION_MARKER)
    a2.font = GRAY_SMALL
    for i, header in enumerate(MANIFEST_HEADERS):
        cell = ws.cell(row=3, column=i + 1, value=header)
        cell.font, cell.fill = Font(bold=True), HEADER_FILL
    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(MANIFEST_HEADERS))}3"


def write_manifest_row(ws, entry: dict) -> None:
    """Upsert the manifest row keyed by sheet name; band + wrap cells."""
    target = None
    for r in range(4, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "") == entry["sheet"]:
            target = r
            break
    if target is None:
        target = 4
        while any(ws.cell(row=target, column=c + 1).value for c in range(len(MANIFEST_HEADERS))):
            target += 1
    values = [entry["sheet"], entry["queryId"], entry["queryName"], entry["sqlSnapshot"], entry["refreshedAt"], entry["rowCount"], entry.get("refreshNotes", "")]
    for i, value in enumerate(values):
        cell = ws.cell(row=target, column=i + 1, value=value)
        cell.alignment = WRAP_TOP
        if target % 2 == 0:
            cell.fill = BAND_FILL
    ws.cell(row=target, column=4).font = MONO_SMALL


def read_csv_file(path: str, sheet: str):
    """Read the caller-written CSV; return (headers, data_rows)."""
    try:
        text = Path(path).read_text(encoding="utf-8-sig")
    except OSError as error:
        fail(2, f"{sheet}: cannot read csvPath {path}: {error}")
    parsed = [row for row in csv.reader(text.splitlines()) if row]
    if not parsed:
        fail(2, f"{sheet}: csvPath {path} is empty")
    return parsed[0], parsed[1:]


# Strict decimal only. int()/float() accept far more than a spreadsheet should
# ("0100"→100, "1_0"→10, "+1"→1, "1e5"→100000.0), and zero-padded GL/cost-center
# codes are the canonical finance identifier — coercing them is silent data
# corruption in the column a user's SUMIFS keys off. So: no leading zeros
# (except a lone 0 integer part), no sign prefix "+", no exponent, no
# underscores, and at most 15 integer digits (Excel's own precision limit).
# Anything that doesn't match stays a string — a string that looks like a
# number is recoverable; a number that used to be an identifier is not.
NUMERIC_PATTERN = re.compile(r"-?(0|[1-9]\d{0,14})(\.\d+)?")


def coerce(value: str):
    """CSV gives strings; store unambiguous decimals as numbers so formulas work."""
    if value == "":
        return None
    if not NUMERIC_PATTERN.fullmatch(value):
        return value
    return float(value) if "." in value else int(value)


def write_data_sheet(wb, entry: dict) -> str:
    headers, rows = read_csv_file(entry["csvPath"], entry["sheet"])
    if len(rows) > ROW_CAP:
        fail(2, f"{entry['sheet']}: {len(rows)} rows exceeds the {ROW_CAP} cap — use a lower grain")
    if entry["sheet"] in wb.sheetnames:
        # Wholesale replace, per contract — but re-insert at the same tab
        # position so refresh never reorders the workbook.
        index = wb.sheetnames.index(entry["sheet"])
        del wb[entry["sheet"]]
        ws = wb.create_sheet(entry["sheet"], index)
    else:
        ws = wb.create_sheet(entry["sheet"])
    ws.sheet_properties.tabColor = TAB_COLOR
    style_band(ws, BANNER_TEMPLATE.format(date=entry["refreshedAt"][:10]), len(headers))
    for i, header in enumerate(headers):
        ws.cell(row=2, column=i + 1, value=header).font = Font(bold=True)
    for r, row in enumerate(rows, start=3):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=coerce(value))
    return f"{entry['sheet']}: {len(rows)} rows · refreshed {entry['refreshedAt']}"


def load_payload(path: str) -> list:
    try:
        with open(path, encoding="utf-8") as f:
            payload = json.load(f)
    except (OSError, json.JSONDecodeError) as error:
        fail(2, f"cannot read payload {path}: {error}")
    entries = payload.get("entries")
    if not isinstance(entries, list) or not entries:
        fail(2, "payload must contain a non-empty 'entries' list")
    required = {"sheet", "queryId", "queryName", "sqlSnapshot", "refreshedAt", "rowCount"}
    for entry in entries:
        missing = required - set(entry)
        if missing:
            fail(2, f"payload entry for {entry.get('sheet', '?')!r} missing {sorted(missing)}")
    return entries


def cmd_list(args) -> None:
    wb = load_workbook(args.workbook, read_only=True)
    _, rows = read_manifest(wb)
    print(json.dumps({"version": VERSION_MARKER, "entries": rows}, indent=2))


def cmd_create(args) -> None:
    entries = load_payload(args.payload)
    try:
        wb = load_workbook(args.workbook)
    except FileNotFoundError:
        wb = Workbook()
        wb.remove(wb.active)
    if MANIFEST_SHEET not in wb.sheetnames:
        style_manifest(wb.create_sheet(MANIFEST_SHEET, 0))
    manifest_ws, _ = read_manifest(wb)
    summaries = []
    for entry in entries:
        if "csvPath" not in entry:
            fail(2, f"create requires csvPath for {entry['sheet']!r}")
        summaries.append(write_data_sheet(wb, entry))
        write_manifest_row(manifest_ws, entry)
    wb.save(args.workbook)
    print("\n".join(summaries))


def cmd_refresh(args) -> None:
    entries = load_payload(args.payload)
    wb = load_workbook(args.workbook)
    manifest_ws, manifest_rows = read_manifest(wb)
    known = {row["sheet"] for row in manifest_rows}
    wanted = {e["sheet"] for e in entries} if args.all or not args.sheets else set(args.sheets.split(","))
    unknown = wanted - known
    if unknown:
        fail(3, f"payload/--sheets name sheets not in the manifest: {sorted(unknown)} — reconcile first")
    missing = [row["sheet"] for row in manifest_rows if row["sheet"] in wanted and row["sheet"] not in wb.sheetnames]
    if missing:
        print(f"note: manifest rows whose sheets were deleted (recreating): {missing}", file=sys.stderr)
    summaries = []
    for entry in entries:
        if entry["sheet"] not in wanted:
            continue
        if "csvPath" not in entry:
            fail(2, f"refresh requires csvPath for {entry['sheet']!r}")
        summaries.append(write_data_sheet(wb, entry))
        write_manifest_row(manifest_ws, entry)
    wb.save(args.workbook)
    print("\n".join(summaries) if summaries else "nothing to refresh")


def main() -> None:
    parser = argparse.ArgumentParser(
        prog="dough_excel.py",
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    sub = parser.add_subparsers(dest="command", required=True)
    for name, handler in (("list", cmd_list), ("create", cmd_create), ("refresh", cmd_refresh)):
        p = sub.add_parser(name)
        p.add_argument("workbook")
        if name != "list":
            p.add_argument("--payload", required=True)
        if name == "refresh":
            p.add_argument("--sheets", default="")
            p.add_argument("--all", action="store_true")
        p.set_defaults(handler=handler)
    args = parser.parse_args()
    args.handler(args)


if __name__ == "__main__":
    main()
